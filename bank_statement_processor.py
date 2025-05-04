import os
import imaplib
import email
import datetime
import re
import pandas as pd
import calendar
import schedule
import time
import PyPDF2
import openpyxl
from email.header import decode_header
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference

class BankStatementProcessor:
    def __init__(self, email_address, password, imap_server, imap_port=993):
        """Initialize with email credentials and server info."""
        self.email_address = email_address
        self.password = password
        self.imap_server = imap_server
        self.imap_port = imap_port
        self.download_folder = "bank_statements"
        self.excel_file = "bank_reports.xlsx"
        
        # Create download folder if it doesn't exist
        if not os.path.exists(self.download_folder):
            os.makedirs(self.download_folder)

    def connect_to_email(self):
        """Connect to the IMAP server."""
        mail = imaplib.IMAP4_SSL(self.imap_server, self.imap_port)
        mail.login(self.email_address, self.password)
        return mail

    def fetch_emails(self, days_back=30):
        """
        Fetch emails that might contain bank statements from the last month.
        """
        print(f"Connecting to {self.imap_server}...")
        mail = self.connect_to_email()
        mail.select('inbox')
        
        # Calculate date for the search (last 30 days by default)
        date = (datetime.datetime.now() - datetime.timedelta(days=days_back)).strftime("%d-%b-%Y")
        
        # Search for emails with subject containing "statement" or "bank"
        search_query = f'(SINCE {date}) SUBJECT "statement" OR SUBJECT "bank"'
        print(f"Searching for emails with query: {search_query}")
        
        status, messages = mail.search(None, search_query)
        email_ids = messages[0].split()
        
        statements = []
        
        for email_id in email_ids:
            status, msg_data = mail.fetch(email_id, '(RFC822)')
            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)
            
            subject = self.decode_email_subject(msg["Subject"])
            sender = msg["From"]
            date_received = msg["Date"]
            
            print(f"Processing email: {subject}")
            
            # Try to identify which bank it's from based on subject or sender
            bank_name = self.identify_bank(subject, sender)
            
            if bank_name:
                # Look for PDF attachments
                for part in msg.walk():
                    if part.get_content_maintype() == 'multipart':
                        continue
                    if part.get('Content-Disposition') is None:
                        continue
                        
                    filename = part.get_filename()
                    if filename and filename.lower().endswith('.pdf'):
                        # Save the attachment
                        filepath = os.path.join(self.download_folder, filename)
                        with open(filepath, 'wb') as f:
                            f.write(part.get_payload(decode=True))
                        
                        print(f"Downloaded: {filename} from {bank_name}")
                        
                        # Add to our list of statements to process
                        statements.append({
                            'bank': bank_name,
                            'filename': filepath,
                            'date': self.parse_date(date_received),
                            'subject': subject
                        })
        
        mail.close()
        mail.logout()
        return statements

    def decode_email_subject(self, subject):
        """Decode email subject properly."""
        if subject is None:
            return ""
        decoded_header = decode_header(subject)
        subject_parts = []
        for content, encoding in decoded_header:
            if isinstance(content, bytes):
                if encoding:
                    content = content.decode(encoding)
                else:
                    # Try utf-8 if no encoding specified
                    try:
                        content = content.decode('utf-8')
                    except UnicodeDecodeError:
                        content = content.decode('latin1', errors='replace')
            subject_parts.append(str(content))
        return ''.join(subject_parts)

    def identify_bank(self, subject, sender):
        """Identify which bank the statement is from based on email subject or sender."""
        subject = subject.lower()
        sender = sender.lower()
        
        # Add your bank identifiers here
        banks = {
            'chase': ['chase', '@chase.com'],
            'bank of america': ['bank of america', 'bankofamerica', '@bofa.com'],
            'wells fargo': ['wells fargo', 'wellsfargo', '@wellsfargo.com'],
            'citi': ['citi', 'citibank', '@citi.com'],
            'capital one': ['capital one', 'capitalone', '@capitalone.com'],
            # Add more banks as needed
        }
        
        for bank, identifiers in banks.items():
            for identifier in identifiers:
                if identifier in subject or identifier in sender:
                    return bank
                    
        # If we can't identify the bank but it contains the word 'statement'
        if 'statement' in subject:
            return 'unknown bank'
            
        return None

    def parse_date(self, date_str):
        """Extract date from email date string."""
        try:
            # Convert email date format to datetime object
            date_tuple = email.utils.parsedate_tz(date_str)
            if date_tuple:
                # Convert to local timestamp
                local_date = datetime.datetime.fromtimestamp(
                    email.utils.mktime_tz(date_tuple))
                return local_date.strftime('%Y-%m')
            return datetime.datetime.now().strftime('%Y-%m')
        except:
            # Default to current month if parsing fails
            return datetime.datetime.now().strftime('%Y-%m')

    def extract_data_from_pdf(self, pdf_path, bank_name):
        """
        Extract financial data from PDF based on bank format.
        Returns a dictionary with extracted data.
        """
        print(f"Extracting data from {pdf_path}...")
        
        data = {
            'bank': bank_name,
            'date': None,
            'closing_balance': None,
            'total_credits': 0,
            'total_debits': 0,
            'statement_period': None
        }
        
        try:
            with open(pdf_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                # Extract text from all pages
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text()
                
                # Different extraction logic based on bank
                if 'chase' in bank_name.lower():
                    data = self.extract_chase_data(text, data)
                elif 'bank of america' in bank_name.lower():
                    data = self.extract_bofa_data(text, data)
                elif 'wells fargo' in bank_name.lower():
                    data = self.extract_wells_fargo_data(text, data)
                elif 'citi' in bank_name.lower():
                    data = self.extract_citi_data(text, data)
                elif 'capital one' in bank_name.lower():
                    data = self.extract_capital_one_data(text, data)
                else:
                    # Generic extraction for unknown banks
                    data = self.extract_generic_bank_data(text, data)
                
                # If we couldn't extract the date, use the filename or current date
                if not data['date']:
                    try:
                        # Try to find a date in the filename
                        date_match = re.search(r'(\d{2,4}[-_]?\d{2}[-_]?\d{2,4})', os.path.basename(pdf_path))
                        if date_match:
                            data['date'] = date_match.group(1)
                        else:
                            # Use the current month as fallback
                            data['date'] = datetime.datetime.now().strftime('%Y-%m')
                    except:
                        data['date'] = datetime.datetime.now().strftime('%Y-%m')
        
        except Exception as e:
            print(f"Error extracting data from PDF: {e}")
            
        return data

    def extract_chase_data(self, text, data):
        """Extract data specific to Chase bank statements."""
        # Try to extract closing balance
        balance_match = re.search(r'[Ee]nding [Bb]alance.*?[\$]?([\d,]+\.\d{2})', text)
        if balance_match:
            data['closing_balance'] = self.parse_amount(balance_match.group(1))
            
        # Try to extract statement period
        period_match = re.search(r'[Ss]tatement [Pp]eriod:?\s*(\d{2}/\d{2}/\d{4})\s*[to|-]\s*(\d{2}/\d{2}/\d{4})', text)
        if period_match:
            data['statement_period'] = f"{period_match.group(1)} - {period_match.group(2)}"
            # Use the end date for the report date
            date_parts = period_match.group(2).split('/')
            data['date'] = f"{date_parts[2]}-{date_parts[0]}"
            
        # Try to extract total deposits/credits
        deposits_match = re.search(r'[Tt]otal [Dd]eposits (?:and [Cc]redits|and [Oo]ther [Aa]dditions).*?[\$]?([\d,]+\.\d{2})', text)
        if deposits_match:
            data['total_credits'] = self.parse_amount(deposits_match.group(1))
            
        # Try to extract total withdrawals/debits
        withdrawals_match = re.search(r'[Tt]otal [Ww]ithdrawals (?:and [Dd]ebits|and [Ff]ees).*?[\$]?([\d,]+\.\d{2})', text)
        if withdrawals_match:
            data['total_debits'] = self.parse_amount(withdrawals_match.group(1))
            
        return data

    def extract_bofa_data(self, text, data):
        """Extract data specific to Bank of America statements."""
        # Try to extract closing balance
        balance_match = re.search(r'[Ee]nding [Bb]alance.*?[\$]?([\d,]+\.\d{2})', text)
        if balance_match:
            data['closing_balance'] = self.parse_amount(balance_match.group(1))
        
        # Try to extract statement period
        period_match = re.search(r'[Ss]tatement [Pp]eriod:?\s*(\d{2}/\d{2}/\d{4})\s*[to|-]\s*(\d{2}/\d{2}/\d{4})', text)
        if period_match:
            data['statement_period'] = f"{period_match.group(1)} - {period_match.group(2)}"
            date_parts = period_match.group(2).split('/')
            data['date'] = f"{date_parts[2]}-{date_parts[0]}"
        
        # Try to extract total deposits
        deposits_match = re.search(r'[Tt]otal deposits.*?[\$]?([\d,]+\.\d{2})', text, re.IGNORECASE)
        if deposits_match:
            data['total_credits'] = self.parse_amount(deposits_match.group(1))
        
        # Try to extract total withdrawals
        withdrawals_match = re.search(r'[Tt]otal withdrawals.*?[\$]?([\d,]+\.\d{2})', text, re.IGNORECASE)
        if withdrawals_match:
            data['total_debits'] = self.parse_amount(withdrawals_match.group(1))
        
        return data

    def extract_wells_fargo_data(self, text, data):
        """Extract data specific to Wells Fargo statements."""
        # Similar pattern as above, customized for Wells Fargo format
        balance_match = re.search(r'[Ee]nding [Bb]alance.*?[\$]?([\d,]+\.\d{2})', text)
        if balance_match:
            data['closing_balance'] = self.parse_amount(balance_match.group(1))
            
        # Statement period
        period_match = re.search(r'[Ss]tatement [Pp]eriod:?\s*(\d{2}/\d{2}/\d{4})\s*[to|-]\s*(\d{2}/\d{2}/\d{4})', text)
        if period_match:
            data['statement_period'] = f"{period_match.group(1)} - {period_match.group(2)}"
            date_parts = period_match.group(2).split('/')
            data['date'] = f"{date_parts[2]}-{date_parts[0]}"
            
        # Deposits
        deposits_match = re.search(r'[Tt]otal [Dd]eposits.*?[\$]?([\d,]+\.\d{2})', text)
        if deposits_match:
            data['total_credits'] = self.parse_amount(deposits_match.group(1))
            
        # Withdrawals
        withdrawals_match = re.search(r'[Tt]otal [Ww]ithdrawals.*?[\$]?([\d,]+\.\d{2})', text)
        if withdrawals_match:
            data['total_debits'] = self.parse_amount(withdrawals_match.group(1))
            
        return data

    def extract_citi_data(self, text, data):
        """Extract data specific to Citibank statements."""
        # Similar pattern as above, customized for Citibank format
        balance_match = re.search(r'[Bb]alance on \d{2}/\d{2}/\d{4}.*?[\$]?([\d,]+\.\d{2})', text)
        if balance_match:
            data['closing_balance'] = self.parse_amount(balance_match.group(1))
            
        # Statement period
        period_match = re.search(r'[Ss]tatement [Pp]eriod:?\s*(\d{2}/\d{2}/\d{4})\s*(?:through|to|-)\s*(\d{2}/\d{2}/\d{4})', text)
        if period_match:
            data['statement_period'] = f"{period_match.group(1)} - {period_match.group(2)}"
            date_parts = period_match.group(2).split('/')
            data['date'] = f"{date_parts[2]}-{date_parts[0]}"
            
        # Credits and debits might need different regex patterns for Citibank
        credits_match = re.search(r'[Tt]otal [Cc]redits.*?[\$]?([\d,]+\.\d{2})', text)
        if credits_match:
            data['total_credits'] = self.parse_amount(credits_match.group(1))
            
        debits_match = re.search(r'[Tt]otal [Dd]ebits.*?[\$]?([\d,]+\.\d{2})', text)
        if debits_match:
            data['total_debits'] = self.parse_amount(debits_match.group(1))
            
        return data

    def extract_capital_one_data(self, text, data):
        """Extract data specific to Capital One statements."""
        # Similar pattern as above, customized for Capital One format
        balance_match = re.search(r'[Ee]nding [Bb]alance.*?[\$]?([\d,]+\.\d{2})', text)
        if balance_match:
            data['closing_balance'] = self.parse_amount(balance_match.group(1))
            
        # Statement period
        period_match = re.search(r'[Ss]tatement [Pp]eriod:?\s*(\d{2}/\d{2}/\d{4})\s*-\s*(\d{2}/\d{2}/\d{4})', text)
        if period_match:
            data['statement_period'] = f"{period_match.group(1)} - {period_match.group(2)}"
            date_parts = period_match.group(2).split('/')
            data['date'] = f"{date_parts[2]}-{date_parts[0]}"
            
        # Credits
        credits_match = re.search(r'[Tt]otal [Cc]redits.*?[\$]?([\d,]+\.\d{2})', text)
        if credits_match:
            data['total_credits'] = self.parse_amount(credits_match.group(1))
            
        # Debits
        debits_match = re.search(r'[Tt]otal [Dd]ebits.*?[\$]?([\d,]+\.\d{2})', text)
        if debits_match:
            data['total_debits'] = self.parse_amount(debits_match.group(1))
            
        return data

    def extract_generic_bank_data(self, text, data):
        """Generic extraction for unknown bank formats."""
        # Try different patterns for closing balance
        balance_patterns = [
            r'[Ee]nding [Bb]alance.*?[\$]?([\d,]+\.\d{2})',
            r'[Bb]alance:?\s*[\$]?([\d,]+\.\d{2})',
            r'[Cc]losing [Bb]alance.*?[\$]?([\d,]+\.\d{2})',
            r'[Tt]otal [Bb]alance.*?[\$]?([\d,]+\.\d{2})'
        ]
        
        for pattern in balance_patterns:
            balance_match = re.search(pattern, text)
            if balance_match:
                data['closing_balance'] = self.parse_amount(balance_match.group(1))
                break
                
        # Try different patterns for statement period
        period_patterns = [
            r'[Ss]tatement [Pp]eriod:?\s*(\d{2}/\d{2}/\d{4})\s*(?:through|to|-)\s*(\d{2}/\d{2}/\d{4})',
            r'[Ss]tatement [Dd]ate:?\s*(\d{2}/\d{2}/\d{4})',
            r'[Pp]eriod:?\s*(\d{2}/\d{2}/\d{4})\s*(?:through|to|-)\s*(\d{2}/\d{2}/\d{4})'
        ]
        
        for pattern in period_patterns:
            period_match = re.search(pattern, text)
            if period_match:
                if len(period_match.groups()) > 1:
                    data['statement_period'] = f"{period_match.group(1)} - {period_match.group(2)}"
                    date_parts = period_match.group(2).split('/')
                else:
                    data['statement_period'] = period_match.group(1)
                    date_parts = period_match.group(1).split('/')
                
                data['date'] = f"{date_parts[2]}-{date_parts[0]}"
                break
                
        # Try different patterns for credits/deposits
        credit_patterns = [
            r'[Tt]otal [Dd]eposits.*?[\$]?([\d,]+\.\d{2})',
            r'[Tt]otal [Cc]redits.*?[\$]?([\d,]+\.\d{2})',
            r'[Dd]eposits [Ss]um:?\s*[\$]?([\d,]+\.\d{2})'
        ]
        
        for pattern in credit_patterns:
            credits_match = re.search(pattern, text)
            if credits_match:
                data['total_credits'] = self.parse_amount(credits_match.group(1))
                break
                
        # Try different patterns for debits/withdrawals
        debit_patterns = [
            r'[Tt]otal [Ww]ithdrawals.*?[\$]?([\d,]+\.\d{2})',
            r'[Tt]otal [Dd]ebits.*?[\$]?([\d,]+\.\d{2})',
            r'[Ww]ithdrawals [Ss]um:?\s*[\$]?([\d,]+\.\d{2})'
        ]
        
        for pattern in debit_patterns:
            debits_match = re.search(pattern, text)
            if debits_match:
                data['total_debits'] = self.parse_amount(debits_match.group(1))
                break
                
        return data

    def parse_amount(self, amount_str):
        """Convert string amount to float."""
        try:
            return float(amount_str.replace(',', ''))
        except:
            return 0.0

    def write_to_excel(self, data_list):
        """Write extracted data to Excel file."""
        print("Writing data to Excel...")
        
        # Create DataFrame from the list of data dictionaries
        df = pd.DataFrame(data_list)
        
        # Calculate net cash flow
        df['net_cash_flow'] = df['total_credits'] - df['total_debits']
        
        # Format date column
        df['month'] = pd.to_datetime(df['date']).dt.strftime('%B %Y')
        
        # Sort by date and bank
        df = df.sort_values(['date', 'bank'])
        
        # Check if file exists, and if not, create a new one
        if not os.path.exists(self.excel_file):
            writer = pd.ExcelWriter(self.excel_file, engine='openpyxl')
            df.to_excel(writer, sheet_name='Bank Statements', index=False)
            writer.close()
        else:
            # Read existing data
            existing_df = pd.read_excel(self.excel_file, sheet_name='Bank Statements')
            
            # Append new data
            combined_df = pd.concat([existing_df, df])
            
            # Remove duplicates (same bank and date)
            combined_df = combined_df.drop_duplicates(subset=['bank', 'date'], keep='last')
            
            # Sort again
            combined_df = combined_df.sort_values(['date', 'bank'])
            
            # Write back to file
            combined_df.to_excel(self.excel_file, sheet_name='Bank Statements', index=False)
            
        return df

    def create_monthly_report(self):
        """Create a monthly report sheet in the Excel file."""
        if not os.path.exists(self.excel_file):
            print("No data available to create monthly report.")
            return
            
        print("Creating monthly report...")
        
        # Load workbook
        workbook = openpyxl.load_workbook(self.excel_file)
        
        # Get the raw data
        statements_sheet = workbook['Bank Statements']
        data = []
        headers = []
        
        # Get headers
        for cell in statements_sheet[1]:
            headers.append(cell.value)
            
        # Get data
        for row in statements_sheet.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            data.append(row_data)
            
        # Convert to DataFrame
        df = pd.DataFrame(data)
        
        # If no data available
        if df.empty:
            print("No data available to create monthly report.")
            workbook.close()
            return
            
        # Get the current month
        current_month = datetime.datetime.now().strftime('%Y-%m')
        
        # Filter data for the current month
        month_data = df[df['date'] == current_month]
        
        if month_data.empty:
            print(f"No data available for {current_month}")
            workbook.close()
            return
            
        # Create monthly report sheet name
        month_year = datetime.datetime.now().strftime('%b_%Y')
        sheet_name = f'Report_{month_year}'
        
        # Check if sheet exists, if yes, remove it
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
            
        # Create new sheet
        report_sheet = workbook.create_sheet(sheet_name)
        
        # Add title
        report_sheet['A1'] = f'Monthly Financial Report - {datetime.datetime.now().strftime("%B %Y")}'
        report_sheet['A1'].font = Font(size=16, bold=True)
        report_sheet.merge_cells('A1:G1')
        report_sheet['A1'].alignment = Alignment(horizontal='center')
        
        # Add summary headers
        report_sheet['A3'] = 'Bank'
        report_sheet['B3'] = 'Closing Balance'
        report_sheet['C3'] = 'Total Credits'
        report_sheet['D3'] = 'Total Debits'
        report_sheet['E3'] = 'Net Cash Flow'
        
        for cell in report_sheet['3:3']:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
        # Add data
        row = 4
        for _, bank_data in month_data.iterrows():
            report_sheet[f'A{row}'] = bank_data['bank']
            report_sheet[f'B{row}'] = bank_data['closing_balance']
            report_sheet[f'C{row}'] = bank_data['total_credits']
            report_sheet[f'D{row}'] = bank_data['total_debits']
            report_sheet[f'E{row}'] = bank_data['net_cash_flow']
            row += 1
            
        # Add totals
        total_row = row
        report_sheet[f'A{total_row}'] = 'TOTAL'
        report_sheet[f'A{total_row}'].font = Font(bold=True)
        
        report_sheet[f'B{total_row}'] = f'=SUM(B4:B{row-1})'
        report_sheet[f'C{total_row}'] = f'=SUM(C4:C{row-1})'
        report_sheet[f'D{total_row}'] = f'=SUM(D4:D{row-1})'
        report_sheet[f'E{total_row}'] = f'=SUM(E4:E{row-1})'
        
        for col in ['B', 'C', 'D', 'E']:
            for r in range(4, row+1):
                report_sheet[f'{col}{r}'].number_format = '$#,##0.00'
                
        # Add a chart
        chart_row = row + 2
        report_sheet[f'A{chart_row}'] = 'Net Cash Flow by Bank'
        report_sheet[f'A{chart_row}'].font = Font(size=12, bold=True)
        
        chart = BarChart()
        chart.title = 'Net Cash Flow by Bank'
        chart.style = 10
        chart.y_axis.title = 'Amount ($)'
        chart.x_axis.title = 'Bank'
        
        data = Reference(report_sheet, min_col=5, min_row=3, max_row=row-1, max_col=5)
        cats = Reference(report_sheet, min_col=1, min_row=4, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        report_sheet.add_chart(chart, f'A{chart_row+2}')
        
        # Adjust column widths
        for column in ['A', 'B', 'C', 'D', 'E']:
            report_sheet.column_dimensions[column].width = 20
            
        # Save the workbook
        workbook.save(self.excel_file)
        workbook.close()
        
        print(f"Monthly report created: {sheet_name}")

    def run_monthly_job(self):
        """Main function to run the monthly job."""
        print(f"Starting monthly bank statement processing: {datetime.datetime.now()}")
        
        # Step 1: Fetch emails and download statements
        statements = self.fetch_emails()
        
        if not statements:
            print("No new bank statements found.")
            return
            
        # Step 2: Extract data from PDFs
        extracted_data = []
        for statement in statements:
            data = self.extract_data_from_pdf(statement['filename'], statement['bank'])
            extracted_data.append(data)
            
        # Step 3: Write data to Excel
        self.write_to_excel(extracted_data)
        
        # Step 4: Create monthly report
        self.create_monthly_report()
        
        print(f"Monthly processing completed: {datetime.datetime.now()}")

def schedule_and_run_job():
    """Set up the scheduler and run the job periodically."""
    # Replace these with your actual email credentials
    email_address = ""
    password = ""  # Use environment variables in production
    imap_server = "imap.gmail.com"  # e.g., imap.gmail.com for Gmail
    
    processor = BankStatementProcessor(email_address, password, imap_server)
    
    # Run immediately for testing
    processor.run_monthly_job()
    
    # Schedule to run on the 1st day of each month at 2:00 AM
    schedule.every().month.at('02:00').do(processor.run_monthly_job)
    
    print("Bank statement processor scheduled to run monthly.")
    print("Press Ctrl+C to exit.")
    
    # Keep the script running
    try:
        while True:
            schedule.run_pending()
            time.sleep(60)  # Check every minute
    except KeyboardInterrupt:
        print("Process terminated by user.")

if __name__ == "__main__":
    schedule_and_run_job()
